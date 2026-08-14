"""Phase 4: Custom roles + permissions, Fee→Ledger, WhatsApp Cloud API, XLSX exports."""
import os, io, httpx
from fastapi import APIRouter, Depends, HTTPException, Request
from fastapi.responses import Response
from pydantic import BaseModel
from typing import List, Optional, Dict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from db import (get_db, get_current_user, require_role, require_school_active,
                new_id, now_utc, iso, audit)

router = APIRouter(prefix="/api", tags=["phase4"])

# ---------------- Permission catalog ----------------
PERMISSION_CATALOG = {
    "students": ["view", "add", "edit", "delete", "export", "print"],
    "admissions": ["view", "add", "edit", "approve", "reject", "convert"],
    "teachers": ["view", "add", "edit", "delete"],
    "parents": ["view", "add", "edit"],
    "attendance": ["view", "add", "edit", "reports"],
    "fees": ["view", "create_invoice", "receive_payment", "edit", "cancel", "print", "reports"],
    "exams": ["view", "create", "enter_marks", "edit_marks", "publish_result", "print_result"],
    "timetable": ["view", "create", "edit", "delete"],
    "expenses": ["view", "add", "edit", "delete", "approve", "reports"],
    "payroll": ["view", "create", "process", "edit", "approve", "print", "reports"],
    "reports": ["view", "export", "print"],
    "settings": ["view", "edit"],
    "notices": ["view", "add", "delete"],
    "diary": ["view", "add", "edit", "delete"],
}

# Default permissions per predefined role (fallback if user has no custom_role_id)
ROLE_DEFAULTS: Dict[str, Dict[str, List[str]]] = {
    "school_admin": {m: acts for m, acts in PERMISSION_CATALOG.items()},
    "accountant": {"fees": ["view","create_invoice","receive_payment","print","reports"],
                   "expenses": ["view","add","edit","reports"],
                   "payroll": ["view","print","reports"],
                   "reports": ["view","export","print"],
                   "students": ["view"], "settings": ["view"]},
    "teacher": {"attendance":["view","add","edit"], "exams":["view","enter_marks","edit_marks","print_result"],
                "diary":["view","add","edit","delete"], "notices":["view","add"],
                "students":["view"], "timetable":["view"]},
    "receptionist": {"admissions": ["view","add","edit","approve","reject","convert"],
                     "students":["view","add","edit"], "parents":["view","add","edit"]},
    "librarian": {"students":["view"]},
    "parent": {}, "student": {}, "principal": {m: acts for m, acts in PERMISSION_CATALOG.items()},
}


async def _resolve_permissions(db, user) -> Dict[str, List[str]]:
    if user.get("custom_role_id"):
        role = await db.custom_roles.find_one({"id": user["custom_role_id"],
                                               "school_id": user.get("school_id")}, {"_id": 0})
        if role:
            return role.get("permissions") or {}
    return ROLE_DEFAULTS.get(user["role"], {})


async def has_permission(db, user, module: str, action: str) -> bool:
    if user["role"] == "super_admin":
        return True
    perms = await _resolve_permissions(db, user)
    return action in perms.get(module, [])


def require_permission(module: str, action: str):
    async def dep(request: Request, user=Depends(get_current_user)):
        db = get_db()
        if not await has_permission(db, user, module, action):
            raise HTTPException(403, f"Missing permission: {module}.{action}")
        return user
    return dep


@router.get("/school/permissions/catalog")
async def catalog(user=Depends(get_current_user)):
    return {"modules": PERMISSION_CATALOG, "role_defaults": ROLE_DEFAULTS}


@router.get("/school/permissions/me")
async def my_perms(user=Depends(get_current_user)):
    db = get_db()
    return {"role": user["role"], "permissions": await _resolve_permissions(db, user)}


# ---------------- Custom roles CRUD ----------------
class CustomRoleIn(BaseModel):
    name: str
    description: Optional[str] = None
    permissions: Dict[str, List[str]] = {}
    active: bool = True


def _sanitize_perms(perms: Dict[str, List[str]]) -> Dict[str, List[str]]:
    """Reject unknown modules/actions and platform-level scopes."""
    clean = {}
    for mod, actions in (perms or {}).items():
        if mod not in PERMISSION_CATALOG:
            continue
        allowed = set(PERMISSION_CATALOG[mod])
        clean[mod] = sorted({a for a in actions if a in allowed})
    return clean


@router.get("/school/custom-roles")
async def list_roles(user=Depends(require_role("school_admin"))):
    db = get_db()
    roles = await db.custom_roles.find({"school_id": user["school_id"]}, {"_id": 0}).to_list(200)
    # count assigned users
    for r in roles:
        r["assigned_users"] = await db.users.count_documents({"custom_role_id": r["id"]})
    return roles


@router.post("/school/custom-roles")
async def create_role(inp: CustomRoleIn, user=Depends(require_role("school_admin"))):
    db = get_db()
    doc = {"id": new_id(), "school_id": user["school_id"], "name": inp.name,
           "description": inp.description, "permissions": _sanitize_perms(inp.permissions),
           "active": inp.active, "created_at": iso(now_utc())}
    await db.custom_roles.insert_one(doc)
    await audit(db, actor=user, action="create_custom_role", module="roles", record_id=doc["id"], after=doc)
    doc.pop("_id", None)
    return doc


@router.put("/school/custom-roles/{rid}")
async def update_role(rid: str, inp: CustomRoleIn, user=Depends(require_role("school_admin"))):
    db = get_db()
    patch = {"name": inp.name, "description": inp.description,
             "permissions": _sanitize_perms(inp.permissions), "active": inp.active}
    r = await db.custom_roles.update_one({"id": rid, "school_id": user["school_id"]}, {"$set": patch})
    if r.matched_count == 0:
        raise HTTPException(404, "Not found")
    await audit(db, actor=user, action="update_custom_role", module="roles", record_id=rid, after=patch)
    return {"ok": True}


@router.delete("/school/custom-roles/{rid}")
async def delete_role(rid: str, user=Depends(require_role("school_admin"))):
    db = get_db()
    assigned = await db.users.count_documents({"custom_role_id": rid, "school_id": user["school_id"]})
    if assigned:
        raise HTTPException(400, f"Role is assigned to {assigned} active users")
    r = await db.custom_roles.delete_one({"id": rid, "school_id": user["school_id"]})
    if r.deleted_count == 0:
        raise HTTPException(404, "Not found")
    await audit(db, actor=user, action="delete_custom_role", module="roles", record_id=rid)
    return {"ok": True}


class AssignRoleIn(BaseModel):
    custom_role_id: Optional[str] = None


@router.put("/school/users/{uid}/custom-role")
async def assign_role(uid: str, inp: AssignRoleIn, user=Depends(require_role("school_admin"))):
    db = get_db()
    target = await db.users.find_one({"id": uid, "school_id": user["school_id"]})
    if not target:
        raise HTTPException(404, "User not found")
    if inp.custom_role_id:
        role = await db.custom_roles.find_one({"id": inp.custom_role_id, "school_id": user["school_id"]})
        if not role:
            raise HTTPException(404, "Role not found (or belongs to another school)")
    await db.users.update_one({"id": uid}, {"$set": {"custom_role_id": inp.custom_role_id}})
    await audit(db, actor=user, action="assign_role", module="roles", record_id=uid,
                after={"custom_role_id": inp.custom_role_id})
    return {"ok": True}


# ---------------- WhatsApp Business Cloud API ----------------
WA_GRAPH_BASE = os.environ.get("WHATSAPP_GRAPH_BASE", "https://graph.facebook.com/v20.0")


def _wa_configured() -> bool:
    return bool(os.environ.get("WHATSAPP_ACCESS_TOKEN") and os.environ.get("WHATSAPP_PHONE_NUMBER_ID"))


async def _wa_send_text(to: str, body: str) -> dict:
    """Real send via Cloud API. Returns dict with provider_message_id or error."""
    if not _wa_configured():
        return {"ok": False, "error": "not_configured"}
    token = os.environ["WHATSAPP_ACCESS_TOKEN"]
    phone_id = os.environ["WHATSAPP_PHONE_NUMBER_ID"]
    url = f"{WA_GRAPH_BASE}/{phone_id}/messages"
    payload = {"messaging_product": "whatsapp", "to": to, "type": "text",
               "text": {"body": body}}
    try:
        async with httpx.AsyncClient(timeout=15) as cli:
            r = await cli.post(url, headers={"Authorization": f"Bearer {token}",
                                             "Content-Type": "application/json"}, json=payload)
        data = r.json() if r.content else {}
        if r.status_code >= 400:
            return {"ok": False, "error": data.get("error", {}).get("message", f"HTTP {r.status_code}"), "raw": data}
        wamid = None
        try: wamid = data.get("messages", [{}])[0].get("id")
        except Exception: pass
        return {"ok": True, "provider_message_id": wamid, "raw": data}
    except Exception as e:
        return {"ok": False, "error": str(e)}


class SendRemindersIn(BaseModel):
    invoice_ids: List[str] = []
    test_mode: bool = False  # true = don't actually call Cloud API even if configured


@router.post("/school/reminders/send-now")
async def send_reminders_now(inp: SendRemindersIn, user=Depends(require_role("school_admin", "accountant"))):
    db = get_db()
    s_id = user["school_id"]
    school = await db.schools.find_one({"id": s_id}, {"_id": 0})
    cfg = school.get("whatsapp_config") or {}
    if not cfg.get("enabled", True):
        raise HTTPException(400, "Reminders are disabled")
    template = cfg.get("template") or "Fee reminder for {student_name}"
    configured = _wa_configured()
    result = {"integration_configured": configured, "test_mode": inp.test_mode, "queued": 0, "sent": 0, "failed": 0}
    for inv_id in inp.invoice_ids:
        inv = await db.fee_invoices.find_one({"id": inv_id, "school_id": s_id}, {"_id": 0})
        if not inv:
            continue
        stu = await db.students.find_one({"id": inv["student_id"], "school_id": s_id}, {"_id": 0}) or {}
        parent = None
        phone = None
        if stu.get("parent_id"):
            parent = await db.users.find_one({"id": stu["parent_id"]}, {"_id": 0, "password_hash": 0})
            phone = (parent or {}).get("phone") or stu.get("phone")
        else:
            phone = stu.get("phone")
        amount_due = inv["amount"] - inv.get("paid_amount", 0)
        message = template.replace("{student_name}", stu.get("name","")).replace(
            "{due_date}", inv["due_date"]).replace("{amount}", f"PKR {amount_due:,.0f}").replace(
            "{school_name}", school.get("name",""))
        log = {"id": new_id(), "school_id": s_id, "invoice_id": inv_id,
               "student_id": inv["student_id"], "parent_id": (parent or {}).get("id"),
               "phone": phone, "message": message, "queued_by": user["id"],
               "integration_configured": configured, "test_mode": inp.test_mode,
               "scheduled_at": iso(now_utc()), "created_at": iso(now_utc())}
        if not configured or inp.test_mode or not phone:
            log["status"] = "queued"
            log["error"] = None if phone else "no_phone"
            result["queued"] += 1
        else:
            log["status"] = "sending"
            res = await _wa_send_text(phone, message)
            if res.get("ok"):
                log["status"] = "sent"
                log["provider_message_id"] = res.get("provider_message_id")
                log["sent_at"] = iso(now_utc())
                result["sent"] += 1
            else:
                log["status"] = "failed"
                log["error"] = res.get("error")
                result["failed"] += 1
        await db.reminder_logs.insert_one(log)
    await audit(db, actor=user, action="send_reminders", module="reminders", after=result)
    if not configured:
        result["integration_note"] = ("WhatsApp Business Cloud API not configured. "
                                      "Messages queued only. Set WHATSAPP_ACCESS_TOKEN and "
                                      "WHATSAPP_PHONE_NUMBER_ID env vars to enable real sending.")
    return result


class WhatsAppStatusIn(BaseModel):
    provider_message_id: str
    status: str  # sent|delivered|read|failed
    error: Optional[str] = None


@router.post("/webhooks/whatsapp/status")
async def wa_status_webhook(inp: WhatsAppStatusIn):
    """Update delivery status. In production, verify webhook signature."""
    db = get_db()
    await db.reminder_logs.update_many(
        {"provider_message_id": inp.provider_message_id},
        {"$set": {"status": inp.status, "error": inp.error, "updated_at": iso(now_utc())}})
    return {"ok": True}


# ---------------- XLSX exports ----------------
def _xlsx_response(wb: Workbook, filename: str) -> Response:
    buf = io.BytesIO(); wb.save(buf)
    return Response(content=buf.getvalue(),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f'attachment; filename="{filename}"'})


HDR_FILL = PatternFill(start_color="065F46", end_color="065F46", fill_type="solid")
HDR_FONT = Font(bold=True, color="FFFFFF")


def _write_header(ws, cols, meta):
    r = 1
    ws.cell(r, 1, meta.get("school_name") or ""); ws.cell(r, 1).font = Font(bold=True, size=14); r += 1
    ws.cell(r, 1, meta.get("report_name") or ""); ws.cell(r, 1).font = Font(bold=True); r += 1
    if meta.get("date_range"): ws.cell(r, 1, meta["date_range"]); r += 1
    ws.cell(r, 1, f"Generated: {now_utc().isoformat()}"); r += 1
    r += 1
    for i, c in enumerate(cols, 1):
        cell = ws.cell(r, i, c); cell.fill = HDR_FILL; cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center")
    return r + 1


@router.get("/school/reports/students.xlsx")
async def students_xlsx(class_id: Optional[str] = None,
                        user=Depends(require_role("school_admin", "accountant"))):
    db = get_db()
    q = {"school_id": user["school_id"]}
    if class_id: q["class_id"] = class_id
    students = await db.students.find(q, {"_id": 0}).sort("name", 1).to_list(20000)
    classes = {c["id"]: c["name"] for c in await db.classes.find({"school_id": user["school_id"]}, {"_id": 0}).to_list(500)}
    school = await db.schools.find_one({"id": user["school_id"]}, {"_id": 0}) or {}
    wb = Workbook(); ws = wb.active; ws.title = "Students"
    cols = ["Student ID", "Adm #", "Name", "Father", "Class", "Roll", "Gender", "DOB", "Phone", "Status"]
    r = _write_header(ws, cols, {"school_name": school.get("name",""), "report_name": "Student List"})
    for s in students:
        ws.append([s.get("student_id",""), s.get("admission_number",""), s["name"],
                   s.get("father_name",""), classes.get(s.get("class_id"),""),
                   s.get("roll_number",""), s.get("gender",""), s.get("dob",""),
                   s.get("phone",""), s.get("status","active")])
    for i in range(1, len(cols)+1): ws.column_dimensions[chr(64+i)].width = 18
    return _xlsx_response(wb, "students.xlsx")


@router.get("/school/reports/fee-collection.xlsx")
async def fee_xlsx(date_from: str, date_to: str,
                   user=Depends(require_role("school_admin", "accountant"))):
    db = get_db()
    s_id = user["school_id"]
    payments = await db.fee_payments.find({"school_id": s_id, "paid_on": {"$gte": date_from, "$lte": date_to}}, {"_id": 0}).sort("paid_on", -1).to_list(20000)
    stu = {s["id"]: s for s in await db.students.find({"school_id": s_id}, {"_id": 0}).to_list(20000)}
    school = await db.schools.find_one({"id": s_id}, {"_id": 0}) or {}
    wb = Workbook(); ws = wb.active; ws.title = "Fee Collection"
    cols = ["Receipt #", "Date", "Student", "Student ID", "Amount", "Method", "Reference"]
    r = _write_header(ws, cols, {"school_name": school.get("name",""),
                                  "report_name": "Fee Collection",
                                  "date_range": f"{date_from} → {date_to}"})
    total = 0
    for p in payments:
        st = stu.get(p.get("student_id"), {})
        ws.append([p.get("receipt_no",""), p.get("paid_on",""), st.get("name",""),
                   st.get("student_id",""), p["amount"], p.get("method",""), p.get("reference","")])
        total += p["amount"]
    row = ws.max_row + 1
    ws.cell(row, 4, "TOTAL").font = Font(bold=True)
    ws.cell(row, 5, total).font = Font(bold=True)
    for i in range(1, len(cols)+1): ws.column_dimensions[chr(64+i)].width = 18
    return _xlsx_response(wb, "fee-collection.xlsx")


@router.get("/school/reports/expenses.xlsx")
async def exp_xlsx(date_from: Optional[str] = None, date_to: Optional[str] = None,
                   user=Depends(require_role("school_admin", "accountant"))):
    db = get_db()
    q = {"school_id": user["school_id"]}
    if date_from or date_to:
        rng = {}
        if date_from: rng["$gte"] = date_from
        if date_to: rng["$lte"] = date_to
        q["date"] = rng
    expenses = await db.expenses.find(q, {"_id": 0}).sort("date", -1).to_list(20000)
    school = await db.schools.find_one({"id": user["school_id"]}, {"_id": 0}) or {}
    wb = Workbook(); ws = wb.active; ws.title = "Expenses"
    cols = ["Date", "Category", "Description", "Amount", "Method", "Status", "Paid To", "Reference", "Created By"]
    r = _write_header(ws, cols, {"school_name": school.get("name",""), "report_name": "Expenses",
                                  "date_range": f"{date_from or ''} → {date_to or ''}"})
    total = 0
    for e in expenses:
        ws.append([e.get("date",""), e.get("category",""), e.get("description",""),
                   e["amount"], e.get("payment_method",""), e.get("status",""),
                   e.get("paid_to",""), e.get("reference",""), e.get("created_by_name","")])
        if e.get("status") == "approved":
            total += e["amount"]
    row = ws.max_row + 1
    ws.cell(row, 3, "APPROVED TOTAL").font = Font(bold=True)
    ws.cell(row, 4, total).font = Font(bold=True)
    for i in range(1, len(cols)+1): ws.column_dimensions[chr(64+i)].width = 18
    return _xlsx_response(wb, "expenses.xlsx")


@router.get("/school/reports/payroll.xlsx")
async def payroll_xlsx(month: str, user=Depends(require_role("school_admin", "accountant"))):
    db = get_db()
    run = await db.payroll_runs.find_one({"school_id": user["school_id"], "month": month}, {"_id": 0})
    if not run:
        raise HTTPException(404, "No payroll run for that month")
    school = await db.schools.find_one({"id": user["school_id"]}, {"_id": 0}) or {}
    wb = Workbook(); ws = wb.active; ws.title = "Payroll"
    cols = ["Employee", "Designation", "Basic", "Allowances", "Deductions", "Net", "Status", "Paid On"]
    r = _write_header(ws, cols, {"school_name": school.get("name",""), "report_name": f"Payroll {month}"})
    for e in run["entries"]:
        ws.append([e["employee_name"], e.get("designation") or "", e["basic"],
                   e["allow_total"], e["deduct_total"], e["net"], e["status"], e.get("paid_on","")])
    row = ws.max_row + 1
    ws.cell(row, 5, "TOTAL NET").font = Font(bold=True)
    ws.cell(row, 6, run["total_net"]).font = Font(bold=True)
    for i in range(1, len(cols)+1): ws.column_dimensions[chr(64+i)].width = 18
    return _xlsx_response(wb, f"payroll-{month}.xlsx")


@router.get("/school/reports/ledger.xlsx")
async def ledger_xlsx(date_from: Optional[str] = None, date_to: Optional[str] = None,
                      user=Depends(require_role("school_admin", "accountant"))):
    db = get_db()
    q = {"school_id": user["school_id"]}
    if date_from or date_to:
        rng = {}
        if date_from: rng["$gte"] = date_from
        if date_to: rng["$lte"] = date_to + "T23:59:59"
        q["created_at"] = rng
    entries = await db.ledger.find(q, {"_id": 0}).sort("created_at", -1).to_list(20000)
    school = await db.schools.find_one({"id": user["school_id"]}, {"_id": 0}) or {}
    wb = Workbook(); ws = wb.active; ws.title = "Ledger"
    cols = ["When", "Kind", "Amount", "Description", "Ref Type", "Ref ID"]
    _write_header(ws, cols, {"school_name": school.get("name",""), "report_name": "Ledger"})
    debit = credit = 0
    for e in entries:
        ws.append([e["created_at"], e["kind"], e["amount"], e.get("description",""),
                   e.get("ref_type",""), e.get("ref_id","")])
        if e["kind"] == "debit": debit += e["amount"]
        else: credit += e["amount"]
    row = ws.max_row + 2
    ws.cell(row, 1, "Totals").font = Font(bold=True)
    ws.cell(row+1, 1, "Credit (Income)"); ws.cell(row+1, 3, credit)
    ws.cell(row+2, 1, "Debit (Expense+Salary)"); ws.cell(row+2, 3, debit)
    ws.cell(row+3, 1, "Net Balance").font = Font(bold=True)
    ws.cell(row+3, 3, credit - debit).font = Font(bold=True)
    for i in range(1, len(cols)+1): ws.column_dimensions[chr(64+i)].width = 22
    return _xlsx_response(wb, "ledger.xlsx")
